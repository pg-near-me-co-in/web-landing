"""
Reads locality/pincode/lat-lng rows out of the India_10_Cities_Pincode_Reference_v2.xlsx
reference file (../data/), one sheet per city. Two things live here:

  load_city_rows(city)   -- every post-office row for a city, as plain dicts
  nearest_row(lat, lng, city_rows) -- which of those rows is geographically
                                       closest to a given point

google_places.py uses both: the first to know which localities to search, the
second to backfill a clean locality/pincode onto a Places result that only
comes with raw coordinates.
"""
import math
import os

import openpyxl

DEFAULT_REFERENCE_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "..", "data",
    "India_10_Cities_Pincode_Reference_v2.xlsx",
)

# Known approximate city-center coordinates, accurate to a few km -- used only
# to catch reference-file rows whose lat/lng doesn't actually belong to their
# stated city. See load_city_rows() docstring for why this exists.
KNOWN_CITY_CENTERS = {
    "Pune": (18.52, 73.86), "Mumbai": (19.08, 72.88), "Bengaluru": (12.97, 77.59),
    "Gurugram": (28.46, 77.03), "Noida": (28.54, 77.39), "Hyderabad": (17.39, 78.49),
    "Ahmedabad": (23.03, 72.58), "Gandhinagar": (23.22, 72.65), "Vadodara": (22.31, 73.19),
    "Kota": (25.21, 75.86),
}
MAX_CITY_RADIUS_KM = 60


def _to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def load_city_rows(city, path=DEFAULT_REFERENCE_FILE):
    """
    Return [{"city", "pincode", "locality", "lat", "lng"}, ...] for every
    post-office row on the given city's sheet. Rows data.gov.in left as "NA"
    for lat/lng are dropped -- useless for locality search or nearest-match.

    A meaningful fraction of rows in this reference file have a lat/lng that
    doesn't actually belong to their stated city -- checked across all 10
    sheets against each city's known center: Mumbai is clean (1% off), but
    Vadodara is 69% off (most of its "bad" rows cluster near Goa instead),
    Kota 34%, Pune 32%, Ahmedabad 19%. Looks like a join/lookup bug from
    whenever this file was generated off the raw data.gov.in export. Rows
    farther than MAX_CITY_RADIUS_KM from the known center are dropped here so
    every consumer (google_places.py, osm_places.py) gets clean anchors --
    this doesn't fix the source file, it just stops a bad row from corrupting
    every query built on top of it. A one-line warning prints when this
    actually drops something, so the problem stays visible instead of silent.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if city not in wb.sheetnames:
        raise ValueError(f"no sheet for city '{city}' in {path} (have: {wb.sheetnames})")
    ws = wb[city]
    center = KNOWN_CITY_CENTERS.get(city)
    rows = []
    dropped = 0
    header = None
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = r
            continue
        row = dict(zip(header, r))
        lat, lng = _to_float(row.get("Latitude")), _to_float(row.get("Longitude"))
        if lat is None or lng is None:
            continue
        if center and haversine_km(center[0], center[1], lat, lng) > MAX_CITY_RADIUS_KM:
            dropped += 1
            continue
        rows.append({
            "city": row.get("Metro_City") or city,
            "pincode": str(row.get("Pincode") or "").strip(),
            "locality": str(row.get("Locality_PostOffice") or "").strip(),
            "lat": lat,
            "lng": lng,
        })
    if dropped:
        print(f"pincode_lookup: dropped {dropped} '{city}' row(s) with a lat/lng "
              f"more than {MAX_CITY_RADIUS_KM}km from {city}'s known center -- "
              f"source reference file data-quality issue, not a bug here.")
    return rows


def haversine_km(lat1, lng1, lat2, lng2):
    r = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lng2 - lng1)
    a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlambda / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


def nearest_city(lat, lng):
    """
    Which KNOWN_CITY_CENTERS key a point is geographically closest to, or None
    if lat/lng is missing. Neighboring pilot cities' search areas overlap
    (Ahmedabad<->Gandhinagar are 22km apart, Gurugram<->Noida 36km) -- a lead
    found in that overlap must be labeled by where it actually is, not by
    whichever city a sweep happened to be targeting, or a later sweep of the
    other city silently relabels (steals) it. See osm_places.py/google_places.py.
    """
    if lat is None or lng is None:
        return None
    return min(KNOWN_CITY_CENTERS, key=lambda c: haversine_km(lat, lng, *KNOWN_CITY_CENTERS[c]))


def nearest_row(lat, lng, city_rows):
    """Closest entry in city_rows to (lat, lng) by straight-line distance, or None if empty."""
    if not city_rows or lat is None or lng is None:
        return None
    return min(city_rows, key=lambda row: haversine_km(lat, lng, row["lat"], row["lng"]))
