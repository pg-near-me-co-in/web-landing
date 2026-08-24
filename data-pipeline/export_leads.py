"""
Export the `leads` collection (osm_places.py / google_places.py output) to an
.xlsx a field agent can work from -- one row per candidate, phone/address/maps
link front and center. Files land in leads_exports/ so they don't clutter the
pipeline directory alongside the scripts.

    python export_leads.py --city Vadodara
    python export_leads.py --city Vadodara --status lead_new   # default: all statuses
    python export_leads.py --city Vadodara --out somewhere/else.xlsx
    python export_leads.py --all-cities                        # one file per pilot city
"""
import argparse
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from db import get_leads_collection
from pincode_lookup import KNOWN_CITY_CENTERS

LEADS_EXPORT_DIR = "leads_exports"

FONT = "Arial"
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
BASE_FONT = Font(name=FONT, size=11)

COLUMNS = [
    "status", "name", "phone", "locality", "pincode", "formatted_address",
    "city", "lat", "lng", "rating", "source", "maps_url",
    "first_seen_at", "last_seen_at", "place_id",
]


def run(city, status=None, out_path=None):
    coll = get_leads_collection()
    query = {"city": city}
    if status:
        query["status"] = status
    docs = list(coll.find(query).sort([("locality", 1), ("name", 1)]))

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    for i, col in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=i, value=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(i)].width = 22

    for r, doc in enumerate(docs, start=2):
        for i, col in enumerate(COLUMNS, start=1):
            v = doc.get(col)
            ws.cell(row=r, column=i, value=str(v) if v is not None else "").font = BASE_FONT

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(len(docs) + 1, 1)}"

    if out_path is None:
        os.makedirs(LEADS_EXPORT_DIR, exist_ok=True)
        out_path = os.path.join(LEADS_EXPORT_DIR, f"leads_{city}.xlsx")
    wb.save(out_path)
    print(f"wrote {len(docs)} lead(s) to {out_path}")
    return out_path


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    group = ap.add_mutually_exclusive_group(required=True)
    group.add_argument("--city")
    group.add_argument("--all-cities", action="store_true", help="write one file per pilot city")
    ap.add_argument("--status", default=None, help="filter to one status (default: all statuses)")
    ap.add_argument("--out", default=None, help="default: leads_exports/leads_<city>.xlsx (ignored with --all-cities)")
    args = ap.parse_args()
    if args.all_cities:
        for c in KNOWN_CITY_CENTERS:
            run(c, args.status)
    else:
        run(args.city, args.status, args.out)
