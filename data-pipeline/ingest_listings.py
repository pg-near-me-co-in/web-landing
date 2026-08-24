"""
Ingest a filled-in intake Excel/CSV (see make_template.py) into MongoDB.

    python ingest_listings.py listings_batch1.xlsx
    python ingest_listings.py listings_batch1.xlsx --batch-size 100

Two independent layers stop duplicates:
  1. A checkpoint file (.checkpoints/<file-hash>.json) remembers which batches
     already finished, so if the process is killed or crashes partway through
     a big file, re-running the exact same command picks up at the next
     unfinished batch instead of starting over.
  2. Every write is an upsert keyed on dedup_key (schema.dedup_key), and that
     field has a UNIQUE index in MongoDB (db.py). So even if a batch somehow
     got re-run, or two people run ingest on overlapping files at the same
     time, Mongo itself refuses to create a second document for the same
     listing — it updates the existing one instead. The checkpoint is for
     speed; the unique index is what actually guarantees no duplicates.

Any row that fails validation is never written to Mongo — it's collected and
dumped to <input>_errors.csv at the end so nothing silently disappears.
"""
import argparse
import csv
import hashlib
import json
import os
import sys
import time
from datetime import datetime

import openpyxl

from db import get_listings_collection
from schema import validate_row, to_mongo_doc, now_utc
from make_template import EXAMPLE_ROW, INTAKE_COLUMNS

CHECKPOINT_DIR = ".checkpoints"


def file_hash(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()[:16]


def read_rows(path):
    if path.lower().endswith((".xlsx", ".xlsm")):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb["Listings"] if "Listings" in wb.sheetnames else wb.active
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for r in ws.iter_rows(min_row=2):
            values = [c.value for c in r]
            if all(v in (None, "") for v in values):
                continue
            rows.append({header[i]: values[i] for i in range(len(header)) if i < len(values)})
        return rows
    else:
        with open(path, newline="", encoding="utf-8") as f:
            return list(csv.DictReader(f))


def is_example_row(row):
    return (
        str(row.get("contact_number", "")).strip() == EXAMPLE_ROW["contact_number"]
        and str(row.get("source_ref", "")).strip() == EXAMPLE_ROW["source_ref"]
    )


def load_checkpoint(path, fhash, batch_size):
    os.makedirs(CHECKPOINT_DIR, exist_ok=True)
    ckpt_path = os.path.join(CHECKPOINT_DIR, f"{os.path.basename(path)}.{fhash}.json")
    if os.path.exists(ckpt_path):
        with open(ckpt_path) as f:
            state = json.load(f)
        if state.get("batch_size") == batch_size:
            return ckpt_path, state
    return ckpt_path, {
        "file": path, "file_hash": fhash, "batch_size": batch_size,
        "last_completed_batch": -1, "inserted": 0, "updated": 0,
        "errors": [], "started_at": now_utc().isoformat(),
    }


def save_checkpoint(ckpt_path, state):
    tmp = ckpt_path + ".tmp"
    with open(tmp, "w") as f:
        json.dump(state, f, indent=2, default=str)
    os.replace(tmp, ckpt_path)  # atomic on POSIX — never leaves a half-written checkpoint


def run(input_path, batch_size=50, simulate_crash_after_batch=None):
    rows = read_rows(input_path)
    rows = [r for r in rows if not is_example_row(r)]
    fhash = file_hash(input_path)
    ckpt_path, state = load_checkpoint(input_path, fhash, batch_size)

    coll = get_listings_collection()
    batches = [rows[i:i + batch_size] for i in range(0, len(rows), batch_size)]
    resumed = state["last_completed_batch"] >= 0
    if resumed:
        print(f"Resuming from checkpoint: batches 0..{state['last_completed_batch']} "
              f"already done, {len(batches) - state['last_completed_batch'] - 1} left.")

    for batch_idx, batch in enumerate(batches):
        if batch_idx <= state["last_completed_batch"]:
            continue  # already committed in a prior run

        for offset, row in enumerate(batch):
            row_num = batch_idx * batch_size + offset + 2  # +2: header row + 1-index
            errors = validate_row(row, row_num)
            if errors:
                state["errors"].append({"row": row_num, "errors": errors, "raw": row})
                continue
            doc = to_mongo_doc(row)
            now = now_utc()
            result = coll.update_one(
                {"dedup_key": doc["dedup_key"]},
                {
                    "$set": {**doc, "updated_at": now},
                    "$setOnInsert": {"created_at": now},
                },
                upsert=True,
            )
            if result.upserted_id is not None:
                state["inserted"] += 1
            elif result.modified_count:
                state["updated"] += 1

        state["last_completed_batch"] = batch_idx
        save_checkpoint(ckpt_path, state)
        print(f"batch {batch_idx + 1}/{len(batches)} committed "
              f"(inserted={state['inserted']} updated={state['updated']} errors={len(state['errors'])})")

        if simulate_crash_after_batch is not None and batch_idx == simulate_crash_after_batch:
            print(f"[TEST HOOK] simulating a crash after batch {batch_idx}")
            raise SystemExit(f"simulated crash after batch {batch_idx}")

    # Full run completed — write the error report (if any) and retire the checkpoint.
    if state["errors"]:
        err_path = os.path.splitext(input_path)[0] + "_errors.csv"
        with open(err_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["row", "errors"] + INTAKE_COLUMNS)
            for e in state["errors"]:
                w.writerow([e["row"], " | ".join(e["errors"])] + [e["raw"].get(c, "") for c in INTAKE_COLUMNS])
        print(f"{len(state['errors'])} invalid rows written to {err_path} — nothing in there reached MongoDB.")

    os.remove(ckpt_path)
    print(f"DONE. inserted={state['inserted']} updated={state['updated']} "
          f"errors={len(state['errors'])} total_rows_seen={len(rows)}")
    return state


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("input", help="xlsx or csv built from the intake template")
    ap.add_argument("--batch-size", type=int, default=50)
    ap.add_argument("--simulate-crash-after-batch", type=int, default=None,
                     help=argparse.SUPPRESS)  # test-only hook, not for normal use
    args = ap.parse_args()
    run(args.input, args.batch_size, args.simulate_crash_after_batch)
