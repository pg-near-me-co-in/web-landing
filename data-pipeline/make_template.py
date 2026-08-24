"""
Generates pg_flat_intake_template.xlsx — the fill-in sheet field agents and
the "list your PG/Flat" form export should use. Feed it straight into
ingest_listings.py.

Run:  python make_template.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from schema import (
    INTAKE_COLUMNS, REQUIRED_COLUMNS, ALLOWED_LISTING_TYPE, ALLOWED_GENDER,
    ALLOWED_AVAILABILITY, ALLOWED_SOURCE,
)

FONT = "Arial"
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
REQUIRED_FILL = PatternFill("solid", fgColor="FFF2CC")
BASE_FONT = Font(name=FONT, size=11)
NOTE_FONT = Font(name=FONT, italic=True, size=10, color="7F7F7F")

FIELD_HELP = {
    "listing_type": "One of: PG, Flat, Room, Shared Room, Independent House",
    "title": "Short headline, e.g. 'Sunshine PG for Working Women, Koramangala'",
    "description": "Free text — anything not covered by the other columns",
    "gender_preference": "Male / Female / Co-ed / Family / Any (blank = Any)",
    "price_monthly": "Monthly rent in INR, numbers only, no commas or ₹",
    "deposit_amount": "Security deposit in INR, 0 if none",
    "availability_status": "Available / Occupied / Coming Soon",
    "available_from": "YYYY-MM-DD, only needed if not already Available",
    "amenities": "Comma-separated, e.g. WiFi, Food, AC, Laundry, Parking, Power Backup",
    "photo_urls": "Comma-separated photo links (upload photos elsewhere first, paste links here)",
    "contact_name": "Owner / manager name",
    "contact_number": "10-digit Indian mobile, no +91 needed",
    "whatsapp_number": "Leave blank to reuse contact_number",
    "address_line": "Building name / street / landmark",
    "locality": "e.g. Koramangala, Andheri West, Gachibowli",
    "city": "One of your pilot cities — Pune, Mumbai, Bengaluru, Gurugram, Noida, Hyderabad, Ahmedabad, Gandhinagar, Vadodara, Kota",
    "state": "e.g. Maharashtra, Karnataka, Telangana",
    "pincode": "6-digit pincode — use the pincode reference file to look it up",
    "latitude": "Leave blank — the ingest script geocodes this from the address automatically",
    "longitude": "Leave blank — the ingest script geocodes this from the address automatically",
    "source": "self_listed / field_agent / partner_broker",
    "source_ref": "Optional — e.g. field agent's name/ID, or a submission form ID",
}

EXAMPLE_ROW = {
    "listing_type": "PG",
    "title": "Sunshine PG for Working Women",
    "description": "3rd floor, lift available, walking distance to metro",
    "gender_preference": "Female",
    "price_monthly": 12000,
    "deposit_amount": 5000,
    "availability_status": "Available",
    "available_from": "",
    "amenities": "WiFi, Food, AC, Laundry, Power Backup",
    "photo_urls": "https://example.com/photo1.jpg, https://example.com/photo2.jpg",
    "contact_name": "Sunita Sharma",
    "contact_number": "9876543210",
    "whatsapp_number": "",
    "address_line": "12, Sunshine Apartments, 5th Cross",
    "locality": "Koramangala",
    "city": "Bengaluru",
    "state": "Karnataka",
    "pincode": "560034",
    "latitude": "",
    "longitude": "",
    "source": "field_agent",
    "source_ref": "agent_ravi_01",
}


def build():
    wb = Workbook()
    wb.remove(wb.active)

    legend = wb.create_sheet("Instructions")
    legend.column_dimensions["A"].width = 22
    legend.column_dimensions["B"].width = 90
    legend["A1"] = "PG / Flat Listing Intake — Instructions"
    legend["A1"].font = Font(name=FONT, bold=True, size=14)
    legend["A2"] = ("Fill one row per listing on the 'Listings' tab. Yellow-shaded columns are required. "
                     "Leave latitude/longitude blank — ingest_listings.py geocodes them from the address. "
                     "Row 2 of the Listings tab is a filled-in example; delete it before you upload a real batch, "
                     "or leave it — the ingest script skips it automatically.")
    legend["A2"].font = BASE_FONT
    legend["A2"].alignment = Alignment(wrap_text=True)
    legend.merge_cells("A2:B2")
    legend.row_dimensions[2].height = 45

    r = 4
    legend.cell(row=r, column=1, value="Column").font = HEADER_FONT
    legend.cell(row=r, column=1).fill = HEADER_FILL
    legend.cell(row=r, column=2, value="What goes here").font = HEADER_FONT
    legend.cell(row=r, column=2).fill = HEADER_FILL
    r += 1
    for col in INTAKE_COLUMNS:
        c1 = legend.cell(row=r, column=1, value=col + (" *" if col in REQUIRED_COLUMNS else ""))
        c1.font = BASE_FONT
        c2 = legend.cell(row=r, column=2, value=FIELD_HELP.get(col, ""))
        c2.font = BASE_FONT
        c2.alignment = Alignment(wrap_text=True)
        r += 1
    legend["A" + str(r + 1)] = "* = required. The ingest script rejects a row (into an errors report) if these are missing or invalid — it never guesses."
    legend["A" + str(r + 1)].font = NOTE_FONT
    legend.merge_cells(f"A{r+1}:B{r+1}")

    ws = wb.create_sheet("Listings")
    for i, col in enumerate(INTAKE_COLUMNS, start=1):
        c = ws.cell(row=1, column=i, value=col)
        c.font = HEADER_FONT
        c.fill = REQUIRED_FILL if col in REQUIRED_COLUMNS else HEADER_FILL
        ws.column_dimensions[get_column_letter(i)].width = 20

    for i, col in enumerate(INTAKE_COLUMNS, start=1):
        ws.cell(row=2, column=i, value=EXAMPLE_ROW.get(col, "")).font = BASE_FONT

    def add_dropdown(col_name, options):
        col_idx = INTAKE_COLUMNS.index(col_name) + 1
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(type="list", formula1=f'"{",".join(sorted(o for o in options if o)) }"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}5000")

    add_dropdown("listing_type", ALLOWED_LISTING_TYPE)
    add_dropdown("gender_preference", ALLOWED_GENDER)
    add_dropdown("availability_status", ALLOWED_AVAILABILITY)
    add_dropdown("source", ALLOWED_SOURCE)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(INTAKE_COLUMNS))}5000"

    wb.move_sheet("Listings", offset=-1)
    wb.save("pg_flat_intake_template.xlsx")
    print("wrote pg_flat_intake_template.xlsx")


if __name__ == "__main__":
    build()
