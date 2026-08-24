"""
Apply many updates at once from a spreadsheet — e.g. a field agent comes back
from re-checking 40 PGs with new prices/availability, or a WhatsApp bot logs
"still available" replies into a CSV you export and feed back in.

Input file needs a "dedup_key" column (or "_id") to say WHICH listing, plus
whichever other columns you want to change — any column left blank is left
untouched on that row, so you only need to fill in what actually changed.

    python bulk_update.py recheck_results.xlsx
    python bulk_update.py recheck_results.xlsx --verify-all

Same checkpoint/resume behaviour as ingest_listings.py — safe to re-run on
the same file after a crash.
"""
import argparse
import csv
import hashlib
import json
import os

import openpyxl
from bson import ObjectId

from db import get_listings_collection
from schema import now_utc, coerce_value

CHECKPOINT_DIR = ".checkpoints"


def file_hash(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()[:16]


def read_rows(path):
    if path.lower().endswith((".xlsx", ".xlsm")):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for r in ws.iter_rows(min_row=2):
            values = [c.value for c in r]
            if all(v in (None, "") for v in values):
                continue
            rows.append({header[i]: values[i] for i in range(len(header)) if i < len(values)})
        return rows
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def run(input_path, batch_size=50, verify_all=False):
    rows = read_rows(input_path)
    fhash = file_hash(input_path)
    os.makedirs(CHECKPOINT_DIR, exist_ok=True)
    ckpt_path = os.path.join(CHECKPOINT_DIR, f"bulkupdate.{os.path.basename(input_path)}.{fhash}.json")
    if os.path.exists(ckpt_path):
        with open(ckpt_path) as f:
            state = json.load(f)
    else:
        state = {"last_completed_batch": -1, "matched": 0, "not_found": []}

    coll = get_listings_collection()
    batches = [rows[i:i + batch_size] for i in range(0, len(rows), batch_size)]

    for batch_idx, batch in enumerate(batches):
        if batch_idx <= state["last_completed_batch"]:
            continue
        for row in batch:
            key = str(row.get("dedup_key", "")).strip()
            _id = str(row.get("_id", "")).strip()
            if not key and not _id:
                state["not_found"].append({"row": row, "reason": "no dedup_key or _id given"})
                continue
            query = {"_id": ObjectId(_id)} if _id else {"dedup_key": key}

            updates = {k: coerce_value(v) for k, v in row.items()
                       if k not in ("dedup_key", "_id") and v not in (None, "")}
            if not updates:
                continue
            updates["updated_at"] = now_utc()
            if verify_all:
                updates["verified"] = True
                updates["last_verified_at"] = now_utc()

            result = coll.update_one(query, {"$set": updates})
            if result.matched_count == 0:
                state["not_found"].append({"row": row, "reason": "no listing matched"})
            else:
                state["matched"] += 1

        state["last_completed_batch"] = batch_idx
        with open(ckpt_path, "w") as f:
            json.dump(state, f, indent=2, default=str)
        print(f"batch {batch_idx + 1}/{len(batches)} committed (matched={state['matched']})")

    if state["not_found"]:
        err_path = os.path.splitext(input_path)[0] + "_not_found.csv"
        with open(err_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["reason", "row"])
            for e in state["not_found"]:
                w.writerow([e["reason"], e["row"]])
        print(f"{len(state['not_found'])} rows didn't match any listing — see {err_path}")

    os.remove(ckpt_path)
    print(f"DONE. matched_and_updated={state['matched']} not_found={len(state['not_found'])}")
    return state


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("input")
    ap.add_argument("--batch-size", type=int, default=50)
    ap.add_argument("--verify-all", action="store_true",
                     help="stamp verified=true + last_verified_at=now on every row in this file")
    args = ap.parse_args()
    run(args.input, args.batch_size, args.verify_all)
